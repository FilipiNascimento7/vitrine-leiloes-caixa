# -*- coding: utf-8 -*-
"""
Coletor de imóveis de leilão da Caixa — Curitiba e Região Metropolitana (RMC).

Baixa a lista pública da Caixa (CSV por estado), filtra por município e
modalidade de venda, e gera:
  - docs/dados.json          -> consumido pela vitrine online (GitHub Pages)
  - docs/vitrine_offline.html-> vitrine com os dados embutidos (abre por duplo clique)
  - docs/imoveis.xlsx        -> planilha para prospecção interna
  - docs/imoveis.csv         -> mesma base em CSV (UTF-8 com BOM, abre no Excel)

Rodar:  python build.py
"""

import csv
import io
import json
import os
import re
import sys
import unicodedata
from datetime import datetime, timezone, timedelta

import requests

# ----------------------------------------------------------------------------
# CONFIGURAÇÃO
# ----------------------------------------------------------------------------

UF = "PR"
CSV_URL = f"https://venda-imoveis.caixa.gov.br/listaweb/Lista_imoveis_{UF}.csv"

# Curitiba + os 28 demais municípios da Região Metropolitana de Curitiba
MUNICIPIOS = [
    "Curitiba",
    "Adrianópolis",
    "Agudos do Sul",
    "Almirante Tamandaré",
    "Araucária",
    "Balsa Nova",
    "Bocaiúva do Sul",
    "Campina Grande do Sul",
    "Campo do Tenente",
    "Campo Largo",
    "Campo Magro",
    "Cerro Azul",
    "Colombo",
    "Contenda",
    "Doutor Ulysses",
    "Fazenda Rio Grande",
    "Itaperuçu",
    "Lapa",
    "Mandirituba",
    "Piên",
    "Pinhais",
    "Piraquara",
    "Quatro Barras",
    "Quitandinha",
    "Rio Branco do Sul",
    "Rio Negro",
    "São José dos Pinhais",
    "Tijucas do Sul",
    "Tunas do Paraná",
]

# Modalidades desejadas. O casamento é por "contém" sobre o texto normalizado,
# então "Venda Direta Online" e "Compra Direta" (nomes que a Caixa alterna) entram.
MODALIDADES_ALVO = [
    "licitacao aberta",
    "venda online",
    "venda direta",
    "compra direta",
]

SAIDA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "docs")
FUSO_BR = timezone(timedelta(hours=-3))

# ----------------------------------------------------------------------------
# UTILITÁRIOS
# ----------------------------------------------------------------------------


def normalizar(texto: str) -> str:
    """minúsculas, sem acento, sem espaços duplicados."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texto))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", t).strip().lower()


MUNICIPIOS_NORM = {normalizar(m): m for m in MUNICIPIOS}


def para_float(valor: str):
    """Aceita os dois formatos que a Caixa mistura no mesmo CSV:
    '1.356.159,46' (br) e '68.18' (ponto decimal, usado na coluna Desconto)."""
    if valor is None:
        return None
    t = re.sub(r"[^\d,.\-]", "", str(valor))
    if not t:
        return None
    if "," in t:                       # vírgula presente => br: ponto é milhar
        t = t.replace(".", "").replace(",", ".")
    elif t.count(".") == 1 and len(t.split(".")[1]) <= 2:
        pass                           # ex.: "68.18" => já é decimal
    else:
        t = t.replace(".", "")         # ex.: "1.356.159" => milhar
    try:
        return float(t)
    except ValueError:
        return None


UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
PAGINA_BUSCA = "https://venda-imoveis.caixa.gov.br/sistema/busca-imovel.asp?sltTipoBusca=imoveis"


def _parece_bloqueio(texto: str) -> bool:
    t = normalizar(texto[:3000])
    return any(x in t for x in ("bot manager", "captcha", "<html", "<head"))


def _baixar_via_requests() -> str:
    resp = requests.get(CSV_URL, timeout=120, headers={"User-Agent": UA})
    resp.raise_for_status()
    return resp.content.decode("latin-1")


class SessaoCaixa:
    """A Caixa bloqueia IPs de datacenter com desafio JS (Radware Bot Manager).
    Abrimos um Chromium real, deixamos o desafio rodar e reaproveitamos a mesma
    sessão para baixar o CSV e depois visitar as páginas de detalhe."""

    def __enter__(self):
        from playwright.sync_api import sync_playwright

        self._pw = sync_playwright().start()
        self.nav = self._pw.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
        )
        ctx = self.nav.new_context(
            user_agent=UA,
            locale="pt-BR",
            timezone_id="America/Sao_Paulo",
            viewport={"width": 1366, "height": 768},
        )
        ctx.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
        )
        self.pg = ctx.new_page()
        self.pg.goto(PAGINA_BUSCA, wait_until="domcontentloaded", timeout=90_000)
        for _ in range(6):                      # espera o desafio JS gravar cookies
            self.pg.wait_for_timeout(3000)
            if not _parece_bloqueio(self.pg.content()[:2000]):
                break
        return self

    def __exit__(self, *exc):
        try:
            self.nav.close()
            self._pw.stop()
        except Exception:
            pass

    def baixar_csv(self) -> str:
        import base64

        b64 = self.pg.evaluate(
            """async (url) => {
                const r = await fetch(url, {credentials: 'include'});
                const buf = new Uint8Array(await r.arrayBuffer());
                let s = '';
                for (const b of buf) s += String.fromCharCode(b);
                return btoa(s);
            }""",
            CSV_URL,
        )
        return base64.b64decode(b64).decode("latin-1")

    def detalhe(self, link: str) -> dict:
        """Lê a página de detalhe do imóvel e devolve texto + PDFs + galeria."""
        self.pg.goto(link, wait_until="domcontentloaded", timeout=60_000)
        self.pg.wait_for_timeout(400)
        return self.pg.evaluate(
            """() => {
                const box = document.querySelector('.content-wrapper.clearfix');
                const texto = box ? box.innerText : '';
                const doc = (re) => {
                    for (const a of document.querySelectorAll('a')) {
                        const oc = a.getAttribute('onclick') || '';
                        const m = oc.match(/ExibeDoc\\('([^']+)'\\)/);
                        if (m && re.test(a.textContent)) return m[1];
                    }
                    return '';
                };
                const fotos = [...new Set(
                    [...document.querySelectorAll('img')]
                      .map(i => i.src)
                      .filter(s => /\\/fotos\\/F\\d+\\.jpg/i.test(s))
                )];
                return {
                    texto,
                    edital:    doc(/edital/i),
                    matricula: doc(/matr/i),
                    fotos,
                };
            }"""
        )


def baixar_csv(sessao=None) -> str:
    print(f"Baixando {CSV_URL} ...")
    try:
        texto = _baixar_via_requests()
        if not _parece_bloqueio(texto):
            print(f"  ok — {len(texto)/1024:.0f} KB (requisição direta)")
            return texto
        print("  bloqueado por desafio anti-bot — usando o navegador…")
    except Exception as e:
        print(f"  requisição direta falhou ({e}) — usando o navegador…")

    if sessao is None:
        raise RuntimeError("Sem sessão de navegador para contornar o anti-bot.")
    texto = sessao.baixar_csv()
    if _parece_bloqueio(texto):
        raise RuntimeError(
            "A Caixa devolveu uma página de desafio anti-bot mesmo pelo navegador."
        )
    print(f"  ok — {len(texto)/1024:.0f} KB (via navegador)")
    return texto


def achar_cabecalho(linhas):
    """A Caixa põe algumas linhas de título antes do cabeçalho real."""
    for i, linha in enumerate(linhas[:15]):
        n = normalizar(linha)
        if "cidade" in n and "modalidade" in n:
            return i
    raise RuntimeError(
        "Não encontrei a linha de cabeçalho no CSV. "
        "O layout da Caixa pode ter mudado — primeiras linhas:\n"
        + "\n".join(linhas[:8])
    )


def achar_coluna(cabecalho, *palavras):
    """Retorna o índice da 1ª coluna cujo nome contém todas as palavras."""
    for i, nome in enumerate(cabecalho):
        n = normalizar(nome)
        if all(p in n for p in palavras):
            return i
    return None


# ----------------------------------------------------------------------------
# PIPELINE
# ----------------------------------------------------------------------------


def coletar(sessao=None):
    bruto = baixar_csv(sessao)
    linhas = [l for l in bruto.splitlines() if l.strip(";").strip()]
    idx_cab = achar_cabecalho(linhas)

    leitor = csv.reader(io.StringIO("\n".join(linhas[idx_cab:])), delimiter=";")
    cabecalho = next(leitor)
    cabecalho = [c.strip() for c in cabecalho]
    print("Colunas detectadas:", cabecalho)

    col = {
        "numero": achar_coluna(cabecalho, "imovel") or 0,
        "uf": achar_coluna(cabecalho, "uf"),
        "cidade": achar_coluna(cabecalho, "cidade"),
        "bairro": achar_coluna(cabecalho, "bairro"),
        "endereco": achar_coluna(cabecalho, "endereco"),
        "preco": achar_coluna(cabecalho, "preco"),
        "avaliacao": achar_coluna(cabecalho, "avaliacao"),
        "desconto": achar_coluna(cabecalho, "desconto"),
        "descricao": achar_coluna(cabecalho, "descricao"),
        "modalidade": achar_coluna(cabecalho, "modalidade"),
        "link": achar_coluna(cabecalho, "link"),
    }
    faltando = [k for k, v in col.items() if v is None and k != "uf"]
    if faltando:
        raise RuntimeError(f"Colunas não encontradas no CSV: {faltando}")

    def campo(linha, chave):
        i = col[chave]
        if i is None or i >= len(linha):
            return ""
        return linha[i].strip()

    imoveis = []
    modalidades_vistas = {}
    total_pr = 0

    for linha in leitor:
        if not linha or len(linha) < 5:
            continue
        total_pr += 1

        cidade_norm = normalizar(campo(linha, "cidade"))
        if cidade_norm not in MUNICIPIOS_NORM:
            continue

        modalidade = campo(linha, "modalidade")
        mod_norm = normalizar(modalidade)
        modalidades_vistas[modalidade] = modalidades_vistas.get(modalidade, 0) + 1
        if not any(alvo in mod_norm for alvo in MODALIDADES_ALVO):
            continue

        numero = campo(linha, "numero")
        preco = para_float(campo(linha, "preco"))
        avaliacao = para_float(campo(linha, "avaliacao"))
        desconto = para_float(campo(linha, "desconto"))

        # desconto real: recalcula quando vier vazio, zerado ou fora de faixa
        if (desconto is None or desconto <= 0 or desconto > 100) and preco and avaliacao and avaliacao > 0:
            desconto = round(max(0.0, (1 - preco / avaliacao) * 100), 2)

        descricao = campo(linha, "descricao")
        link = campo(linha, "link")

        # foto: padrão do portal é F + número preenchido com zeros até 13 dígitos + "21.jpg"
        # ex.: imóvel 10139954 -> F000001013995421.jpg
        m = re.search(r"hdnimovel=(\d+)", link)
        hdn = m.group(1) if m else re.sub(r"\D", "", numero)
        foto = (
            f"https://venda-imoveis.caixa.gov.br/fotos/F{hdn.zfill(13)}21.jpg"
            if hdn else ""
        )

        imoveis.append(
            {
                "numero": numero,
                "cidade": MUNICIPIOS_NORM[cidade_norm],
                "bairro": campo(linha, "bairro").title(),
                "endereco": campo(linha, "endereco"),
                "preco": preco,
                "avaliacao": avaliacao,
                "desconto": desconto,
                "economia": (round(avaliacao - preco, 2) if (preco and avaliacao) else None),
                "tipo": classificar_tipo(descricao),
                "area": extrair_area(descricao),
                "quartos": extrair_quartos(descricao),
                "aceita_financiamento": "financiamento" in normalizar(descricao)
                and "nao aceita financiamento" not in normalizar(descricao),
                "aceita_fgts": "fgts" in normalizar(descricao)
                and "nao aceita fgts" not in normalizar(descricao),
                "descricao": descricao,
                "modalidade": modalidade,
                "link": link,
                "foto": foto,
            }
        )

    imoveis.sort(key=lambda i: (i["desconto"] or 0), reverse=True)
    return imoveis, total_pr, modalidades_vistas


# ----------------------------------------------------------------------------
# DETALHES (uma visita por imóvel na página oficial)
# ----------------------------------------------------------------------------

BASE = "https://venda-imoveis.caixa.gov.br"

# blocos de texto livre que a Caixa exibe na página de detalhe
BLOCOS = [
    ("descricao_completa", "Descrição:"),
    ("formas_pagamento", "FORMAS DE PAGAMENTO ACEITAS:"),
    ("regras_despesas", "REGRAS PARA PAGAMENTO DAS DESPESAS"),
]


def parsear_detalhe(texto: str) -> dict:
    """Transforma o texto da página de detalhe em campos + blocos."""
    # a ficha (pares "Chave: valor") só existe antes do bloco "Descrição:" — depois
    # dele vêm textos livres que também têm dois-pontos e não são campos.
    corte = texto.find("Descrição:")
    cabeca = texto[:corte] if corte > 0 else texto

    linhas = [l.strip() for l in cabeca.splitlines() if l.strip()]
    campos, areas = {}, {}
    endereco_completo = ""

    for i, l in enumerate(linhas):
        # "Área privativa = 2.890,80m2"
        m = re.match(r"^(Área[^=]+)=\s*([\d.,]+)\s*m2?$", l, re.I)
        if m:
            areas[m.group(1).strip()] = m.group(2).strip() + " m²"
            continue
        # "Chave: valor"
        m = re.match(r"^([A-Za-zÀ-ú()\s/º°.-]{3,45}?):\s*(.+)$", l)
        if m and not l.upper().startswith(("FORMAS", "REGRAS", "DESCRIÇÃO")):
            chave, valor = m.group(1).strip(), m.group(2).strip()
            if chave.lower() == "endereço":
                continue
            campos.setdefault(chave, valor)
            continue
        if l.lower().startswith("endereço"):
            endereco_completo = linhas[i + 1] if i + 1 < len(linhas) else ""
        if re.search(r"data d[ao] (licitação|leilão|venda)", l, re.I):
            campos.setdefault("Data da disputa", l.lstrip("• ").strip())

    # blocos de texto livre — do rótulo até o próximo rótulo conhecido
    rotulos = [r for _, r in BLOCOS] + ["Baixar", "Voltar", "Dê seu lance"]
    blocos = {}
    for chave, rotulo in BLOCOS:
        m = re.search(re.escape(rotulo) + r"[^\n]*\n", texto)
        if not m:
            continue
        resto = texto[m.end():]
        fim = len(resto)
        for r in rotulos:
            if r == rotulo:
                continue
            p = resto.find(r)
            if 0 <= p < fim:
                fim = p
        bruto_bloco = resto[:fim].strip()
        conteudo = "".join(c if (c.isprintable() or c == chr(10)) else " " for c in bruto_bloco)
        if conteudo:
            blocos[chave] = [x.strip(" •·\t") for x in conteudo.splitlines() if x.strip()]

    return {
        "campos": campos,
        "areas": areas,
        "blocos": blocos,
        "endereco_completo": endereco_completo,
    }


def enriquecer(imoveis, sessao):
    """Visita a página oficial de cada imóvel e guarda TUDO que a Caixa mostra."""
    print(f"\nColetando os detalhes de {len(imoveis)} imóveis (página por página)…")
    ok = falhas = 0
    for n, im in enumerate(imoveis, 1):
        try:
            d = sessao.detalhe(im["link"])
            det = parsear_detalhe(d.get("texto", ""))
            det["fotos"] = d.get("fotos") or ([im["foto"]] if im["foto"] else [])
            det["edital"] = (BASE + d["edital"]) if d.get("edital") else ""
            det["matricula"] = (BASE + d["matricula"]) if d.get("matricula") else ""
            im["detalhe"] = det
            if det["fotos"]:
                im["foto"] = det["fotos"][0]
            im["pagina"] = f"imovel/{im['numero']}.html"
            ok += 1
        except Exception as e:
            falhas += 1
            im["detalhe"] = None
            im["pagina"] = ""
            print(f"  ! falhou no imóvel {im['numero']}: {e}")
        if n % 25 == 0 or n == len(imoveis):
            print(f"  {n}/{len(imoveis)}…")
    print(f"  detalhes coletados: {ok} ok, {falhas} falhas")


def classificar_tipo(descricao: str) -> str:
    d = normalizar(descricao)
    for chave, rotulo in [
        ("apartamento", "Apartamento"),
        ("casa", "Casa"),
        ("sobrado", "Casa"),
        ("terreno", "Terreno"),
        ("lote", "Terreno"),
        ("gleba", "Terreno"),
        ("comercial", "Comercial"),
        ("loja", "Comercial"),
        ("sala", "Comercial"),
        ("galpao", "Comercial"),
        ("predio", "Comercial"),
        ("rural", "Rural"),
        ("chacara", "Rural"),
        ("fazenda", "Rural"),
    ]:
        if chave in d:
            return rotulo
    return "Outros"


def extrair_area(descricao: str):
    """A Caixa lista várias áreas ('0.00 de área total, 72600.00 de área do terreno').
    Preferimos privativa > total > terreno, ignorando zeros."""
    d = normalizar(descricao)
    achados = {}
    for valor, rotulo in re.findall(
        r"([\d.,]+)\s*(?:m2\s*)?de area (privativa|total|do terreno|util)", d
    ):
        v = para_float(valor)
        if v and v > 0 and rotulo not in achados:
            achados[rotulo] = v
    for rotulo in ("privativa", "util", "total", "do terreno"):
        if rotulo in achados:
            return achados[rotulo]
    m = re.search(r"([\d.,]+)\s*(?:m2|m²)", d)
    v = para_float(m.group(1)) if m else None
    return v if (v and v > 0) else None


def extrair_quartos(descricao: str):
    d = normalizar(descricao)
    m = re.search(r"(\d+)\s*(?:quarto|dormitorio|qto|dorm)", d)
    return int(m.group(1)) if m else None


# ----------------------------------------------------------------------------
# SAÍDAS
# ----------------------------------------------------------------------------


def gravar(imoveis, total_pr, modalidades_vistas):
    os.makedirs(SAIDA, exist_ok=True)
    agora = datetime.now(FUSO_BR)

    resumo = {
        "atualizado_em": agora.strftime("%d/%m/%Y às %H:%M"),
        "atualizado_iso": agora.isoformat(),
        "total": len(imoveis),
        "total_pr": total_pr,
        "municipios": sorted({i["cidade"] for i in imoveis}),
        "modalidades": sorted({i["modalidade"] for i in imoveis}),
        "valor_total": round(sum(i["preco"] or 0 for i in imoveis), 2),
        "desconto_medio": (
            round(sum(i["desconto"] or 0 for i in imoveis) / len(imoveis), 1) if imoveis else 0
        ),
    }

    # 1) páginas de detalhe (uma por imóvel, com galeria e PDFs)
    import pagina_imovel
    pagina_imovel.gravar_todas(imoveis, resumo["atualizado_em"], SAIDA)

    # o dados.json da vitrine fica enxuto: o detalhe pesado vive na página do imóvel
    leves = []
    for im in imoveis:
        det = im.get("detalhe") or {}
        leve = {k: v for k, v in im.items() if k != "detalhe"}
        leve["edital"] = det.get("edital", "")
        leve["matricula"] = det.get("matricula", "")
        leve["n_fotos"] = len(det.get("fotos") or [])
        leves.append(leve)

    with open(os.path.join(SAIDA, "dados.json"), "w", encoding="utf-8") as f:
        json.dump({"resumo": resumo, "imoveis": leves}, f, ensure_ascii=False, indent=1)
    imoveis = leves

    # 2) CSV (Excel-friendly)
    campos = [
        "numero", "cidade", "bairro", "endereco", "tipo", "area", "quartos",
        "preco", "avaliacao", "desconto", "economia", "modalidade",
        "aceita_financiamento", "aceita_fgts", "link", "edital", "matricula", "descricao",
    ]
    with open(os.path.join(SAIDA, "imoveis.csv"), "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=campos, delimiter=";", extrasaction="ignore")
        w.writeheader()
        w.writerows(imoveis)

    # 3) XLSX (prospecção interna) — opcional, só se openpyxl estiver instalado
    try:
        gravar_xlsx(imoveis, resumo, campos)
    except ImportError:
        print("  (openpyxl não instalado — pulei o .xlsx; o .csv foi gerado)")

    # 4) Vitrine offline com dados embutidos
    modelo_path = os.path.join(SAIDA, "index.html")
    if os.path.exists(modelo_path):
        with open(modelo_path, encoding="utf-8") as f:
            html = f.read()
        embutido = (
            "<script>window.DADOS_EMBUTIDOS = "
            + json.dumps({"resumo": resumo, "imoveis": imoveis}, ensure_ascii=False)
            + ";</script>"
        )
        html_off = html.replace("<!--DADOS-->", embutido)
        with open(os.path.join(SAIDA, "vitrine_offline.html"), "w", encoding="utf-8") as f:
            f.write(html_off)

    print("\n" + "=" * 62)
    print(f"  {resumo['total']} imóveis em Curitiba + RMC "
          f"(de {total_pr} no PR inteiro)")
    print(f"  Desconto médio: {resumo['desconto_medio']}%")
    print(f"  Atualizado em:  {resumo['atualizado_em']}")
    print("=" * 62)
    print("\n  Modalidades encontradas no PR (as marcadas com * entraram):")
    for mod, qtd in sorted(modalidades_vistas.items(), key=lambda x: -x[1]):
        marca = "*" if any(a in normalizar(mod) for a in MODALIDADES_ALVO) else " "
        print(f"   {marca} {mod or '(vazio)'}: {qtd}")
    print(f"\n  Arquivos gerados em: {SAIDA}\n")


def gravar_xlsx(imoveis, resumo, campos):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Imóveis"

    titulos = {
        "numero": "Nº Imóvel", "cidade": "Cidade", "bairro": "Bairro",
        "endereco": "Endereço", "tipo": "Tipo", "area": "Área (m²)",
        "quartos": "Quartos", "preco": "Preço (R$)", "avaliacao": "Avaliação (R$)",
        "desconto": "Desconto (%)", "economia": "Economia (R$)",
        "modalidade": "Modalidade", "aceita_financiamento": "Financia?",
        "aceita_fgts": "FGTS?", "link": "Link Caixa", "edital": "Edital (PDF)",
        "matricula": "Matrícula (PDF)", "descricao": "Descrição",
    }

    ws.append([titulos[c] for c in campos])
    cab_fill = PatternFill("solid", fgColor="1B4D8F")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = cab_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for im in imoveis:
        ws.append([im.get(c) for c in campos])

    for i, c in enumerate(campos, start=1):
        letra = get_column_letter(i)
        larguras = {"endereco": 42, "descricao": 60, "link": 30, "bairro": 20,
                    "cidade": 20, "modalidade": 20}
        ws.column_dimensions[letra].width = larguras.get(c, 14)
        if c in ("preco", "avaliacao", "economia"):
            for cel in ws[letra][1:]:
                cel.number_format = 'R$ #,##0.00'
        if c == "desconto":
            for cel in ws[letra][1:]:
                cel.number_format = '0.0"%"'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Resumo")
    ws2.append(["Atualizado em", resumo["atualizado_em"]])
    ws2.append(["Imóveis (Curitiba + RMC)", resumo["total"]])
    ws2.append(["Imóveis no PR inteiro", resumo["total_pr"]])
    ws2.append(["Desconto médio (%)", resumo["desconto_medio"]])
    ws2.append(["Valor total ofertado (R$)", resumo["valor_total"]])
    ws2.append([])
    ws2.append(["Modalidades incluídas"])
    for m in resumo["modalidades"]:
        ws2.append(["", m])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 40
    for c in ("A1", "A2", "A3", "A4", "A5", "A7"):
        ws2[c].font = Font(bold=True)

    wb.save(os.path.join(SAIDA, "imoveis.xlsx"))


if __name__ == "__main__":
    try:
        with SessaoCaixa() as sessao:
            imoveis, total_pr, mods = coletar(sessao)
            if imoveis:
                enriquecer(imoveis, sessao)
    except Exception as e:
        print(f"\nERRO: {e}\n", file=sys.stderr)
        sys.exit(1)
    if not imoveis:
        print("\nAVISO: nenhum imóvel bateu com os filtros. "
              "Confira as modalidades listadas abaixo e ajuste MODALIDADES_ALVO.\n")
    gravar(imoveis, total_pr, mods)
